from openpyxl import load_workbook


class XlsxParser:
    EXTENSIONS = {'.xlsx'}
    OUTPUT_EXTENSION = ".xlsx"

    def parse(self, file_path: str) -> str:
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception:
            return ""
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        parts.append(str(cell))
        return '\n'.join(parts)

    def anonymize_to_path(self, file_path: str, output_path: str, anonymizer):
        try:
            wb = load_workbook(file_path)
        except Exception:
            return False, []

        spans = []
        changed = False

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    if cell.data_type == "f":
                        continue

                    anonymized, cell_spans = anonymizer._anonymize_text_with_spans(cell.value)
                    if cell_spans:
                        cell.value = anonymized
                        spans.extend(cell_spans)
                        changed = True

        if not changed:
            return False, []

        wb.save(output_path)
        return True, spans
